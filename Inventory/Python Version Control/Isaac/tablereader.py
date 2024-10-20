from PySide6.QtWidgets import QTableWidgetItem
import openpyxl

# Exports a table widget's data to openpyxl workbook
def export_table(table):
    # Create workbook and select active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Read table
    (headers, data) = read_table(table)

    # Write headers to workbook
    ws.append(headers)

    # Write rows to workbook
    for row in data:
        ws.append(row)

    return wb

# Reads a table widget's data into a tuple in format ([headers], [data])
def read_table(table):
    headers = []
    # Get header names
    for column in range(table.columnCount()):
        if table.horizontalHeaderItem(column) is None:
            table.setHorizontalHeaderItem(column, QTableWidgetItem(""))
        headers.append(table.horizontalHeaderItem(column).text())

    # Get table data
    data = []
    for row in range(table.rowCount()):
        row_data = []
        for column in range(table.columnCount()):
            if table.item(row, column) is None:
                table.setItem(row, column, QTableWidgetItem(""))
            row_data.append(table.item(row, column).text())

        data.append(row_data)

    return headers, data


#TEST DATA
test_headers = ['Name', 'Category', 'Quantity', 'Cost', 'Sale Price', 'Available', 'Date Stocked', 'Contact']
test_rows = [
    ["Alice's Apples", 'Fruit', 50, 1.50, 3.00, 'Yes', '2024-09-12', 'alice@example.com'],
    ["Bob's Books", 'Literature', 30, 12.00, 18.00, 'No', '2023-11-01', 'bob@example.com'],
    ["Charlie's Chairs", 'Furniture', 10, 45.00, 80.00, 'Yes', '2024-06-22', 'charlie@example.com'],
    ["Dan's Drills", 'Tools', 20, 25.00, 40.00, 'Yes', '2024-02-15', 'dan@example.com'],
    ["Eve's Earrings", 'Jewelry', 100, 5.00, 12.00, 'Yes', '2024-01-10', 'eve@example.com'],
    ["Frank's Frames", 'Art Supplies', 15, 8.00, 15.00, 'Yes', '2024-05-03', 'frank@example.com'],
    ["Gina's Games", 'Toys', 40, 10.00, 20.00, 'No', '2024-07-12', 'gina@example.com'],
    ["Harry's Hats", 'Apparel', 25, 7.00, 14.00, 'Yes', '2024-03-29', 'harry@example.com'],
    ["Isla's Instruments", 'Music', 5, 100.00, 180.00, 'No', '2024-08-18', 'isla@example.com'],
    ["Jack's Jackets", 'Apparel', 12, 35.00, 60.00, 'Yes', '2024-04-20', 'jack@example.com']
]
test_headers2 = []
test_rows2 = []

# Imports data from an openpyxl workbook into a tuple in format ([headers], [data])
def import_workbook():
    # USE FUNCTION FROM EXCEL HANDLING TO GET WORKBOOK
    # Temp workbook
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.append(test_headers)
    for row in test_rows:
        ws.append(row)

    rows = ws.iter_rows(values_only=True)
    headers = [header for header in next(rows)] # Get header names

    # Get table data
    data = []
    for row in rows:
        data.append(row)

    return headers, data