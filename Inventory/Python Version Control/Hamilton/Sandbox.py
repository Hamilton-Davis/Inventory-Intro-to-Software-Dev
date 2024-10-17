import datetime
import xlwt
import os
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from pathlib import Path
from dateutil.parser import parse

#Gets path to most recent .xlsx file in directory
def get_newest_file(folderPath):
    files = []
    for filePath in Path.iterdir(folderPath):
        if filePath.suffix == '.xlsx':
            try:
                date = parse(filePath.stem, fuzzy=True) # Pulls first 3-integer date (e.g. xx-xx-xxxx) found in file name (stem)
                file = (date, filePath)
                files.append(file)
            except ValueError: # Could not parse 3-integer date from file name
                pass

    files.sort(reverse=True) # Sort in descending order, newest dates first
    return files[0][1] # Returns filePath of newest file


#folderLocation = "C:\\Users\\mirac\\OneDrive\\Desktop\\Inventory-Intro-to-Software-Dev-main\\Inventory-Intro-to-Software-Dev-main\\Inventory Files\\Excel Sheets\\Excel Sheet"
folderLocation = Path(__file__).parent.parent.parent.absolute() / "Excel Sheets" # Absolute path to Inventory folder:
# Goes up from script file location to parent "Hamilton", parent "Python Version Control", parent "Inventory", appends "Excel Sheets"


# Finds Current Date, Allows for File Management
current_time = datetime.datetime.now()
month = str(current_time.month) + '-'
day = str(current_time.day) + '-'
year = str(current_time.year)
date = month+day+year
# print(date)

# Defines file location for new file
excelName = "Inventory-"+ date + ".xlsx"
newFileLocation = folderLocation / excelName

print(newFileLocation)

wb = load_workbook(get_newest_file(folderLocation)) # Loads most recent Excel sheet into workbook
sheet = wb.active # Active sheet in workbook (IDK if this part is necessary)
for row in sheet.iter_rows(values_only=True):
    print(row)

"""# Default File, It must have this on the file
for filename in Path.iterdir(folderLocation):
    if filename.suffix == ".xlsx":
        copyPath = folderLocation / filename
        print(filename)
    else:
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Category"
        ws['B1'] = "Item"
        ws['C1'] = "InStock"
        ws['D1'] = "Purchased"
        ws['E1'] = "Price"
        ws['F1'] = "Availability"
        ws['G1'] = "Link"
        ws['H1'] = "Entry Date"
"""
#wb.save(filename) # This saves file to current directory of script file, not to newFileLocation
wb.save(newFileLocation)
